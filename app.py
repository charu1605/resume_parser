from flask import Flask, request, render_template
import pickle
from PyPDF2 import PdfReader
import re
import os
import sqlite3

# ===================== DATABASE CONFIG =====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "resume.db")

import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "resume_data.xlsx"

def update_excel(data_dict):
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Parsed Resumes"
            headers = ["Name", "Email", "Phone", "Category", "Job Recommendation", "Skills", "Education"]
            ws.append(headers)
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        row = [
            data_dict.get("name", "N/A"),
            data_dict.get("email", "N/A"),
            data_dict.get("phone", "N/A"),
            data_dict.get("category", "N/A"),
            data_dict.get("job", "N/A"),
            ", ".join(data_dict.get("skills", [])),
            ", ".join(data_dict.get("education", []))
        ]
        ws.append(row)
        wb.save(EXCEL_FILE)
    except PermissionError:
        print("Error: Could not save to Excel. Please close 'resume_data.xlsx' and try again.")

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS resumes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            email TEXT,
            phone TEXT,
            category TEXT,
            recommended_job TEXT,
            skills TEXT,
            education TEXT
        )
    """)

    conn.commit()
    conn.close()

def save_resume_to_db(name, email, phone, category, recommended_job, skills, education):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO resumes 
        (name, email, phone, category, recommended_job, skills, education)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        name,
        email,
        phone,
        category,
        recommended_job,
        ", ".join(skills),
        ", ".join(education)
    ))

    conn.commit()
    conn.close()

# ===================== FLASK APP =====================
app = Flask(__name__)

# Use absolute paths to avoid "File Not Found" errors
MODELS_DIR = os.path.join(BASE_DIR, 'models')

rf_classifier_categorization = pickle.load(open(os.path.join(MODELS_DIR, 'rf_classifier_categorization.pkl'), 'rb'))
tfidf_vectorizer_categorization = pickle.load(open(os.path.join(MODELS_DIR, 'tfidf_vectorizer_categorization.pkl'), 'rb'))
rf_classifier_job_recommendation = pickle.load(open(os.path.join(MODELS_DIR, 'rf_classifier_job_recommendation.pkl'), 'rb'))
tfidf_vectorizer_job_recommendation = pickle.load(open(os.path.join(MODELS_DIR, 'tfidf_vectorizer_job_recommendation.pkl'), 'rb'))
# ===================== HELPERS =====================
import pdfplumber

def pdf_to_text(file):
    with pdfplumber.open(file) as pdf:
        # pdfplumber is better at identifying spaces between layout elements
        return " ".join(page.extract_text() or "" for page in pdf.pages)

def cleanResume(txt):
    txt = re.sub(r'http\S+|RT|cc|@\S+|#\S+', ' ', txt)
    txt = re.sub(r'[^\x00-\x7f]', ' ', txt)
    txt = re.sub(r'\s+', ' ', txt)
    return txt

def predict_category(text):
    vec = tfidf_vectorizer_categorization.transform([text])
    return rf_classifier_categorization.predict(vec)[0]

def job_recommendation(text):
    vec = tfidf_vectorizer_job_recommendation.transform([text])
    return rf_classifier_job_recommendation.predict(vec)[0]

def extract_email_from_resume(text):
    match = re.search(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
    return match.group() if match else None

def extract_contact_number_from_resume(text):
    match = re.search(r"\b(?:\+?\d{1,3})?[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}\b", text)
    return match.group() if match else None


def extract_name_from_resume(text):
    # Only look at the header (first 150 chars) where names usually live
    header = text[:150].strip()

    # This pattern matches "Charu Garg" OR "CHARU GARG"
    # It requires word boundaries and looks for 2-3 capitalized words
    name_pattern = r"\b([A-Z][a-z]+|[A-Z]{2,})\s+([A-Z][a-z]+|[A-Z]{2,})\b"

    match = re.search(name_pattern, header)
    if match:
        name = match.group()
        # Filter out common false positives found in headers
        forbidden = ["Resume", "Curriculum", "Profile", "Objective"]
        if any(word in name for word in forbidden):
            return None
        return name
    return None

def extract_skills_from_resume(text):
    # List of predefined skills
    skills_list = [
        'Python', 'Data Analysis', 'Machine Learning', 'Communication', 'Project Management', 'Deep Learning', 'SQL',
        'Tableau',
        'Java', 'C++', 'JavaScript', 'HTML', 'CSS', 'React', 'Angular', 'Node.js', 'MongoDB', 'Express.js', 'Git',
        'Research', 'Statistics', 'Quantitative Analysis', 'Qualitative Analysis', 'SPSS', 'R', 'Data Visualization',
        'Matplotlib',
        'Seaborn', 'Plotly', 'Pandas', 'Numpy', 'Scikit-learn', 'TensorFlow', 'Keras', 'PyTorch', 'NLTK', 'Text Mining',
        'Natural Language Processing', 'Computer Vision', 'Image Processing', 'OCR', 'Speech Recognition',
        'Recommendation Systems',
        'Collaborative Filtering', 'Content-Based Filtering', 'Reinforcement Learning', 'Neural Networks',
        'Convolutional Neural Networks',
        'Recurrent Neural Networks', 'Generative Adversarial Networks', 'XGBoost', 'Random Forest', 'Decision Trees',
        'Support Vector Machines',
        'Linear Regression', 'Logistic Regression', 'K-Means Clustering', 'Hierarchical Clustering', 'DBSCAN',
        'Association Rule Learning',
        'Apache Hadoop', 'Apache Spark', 'MapReduce', 'Hive', 'HBase', 'Apache Kafka', 'Data Warehousing', 'ETL',
        'Big Data Analytics',
        'Cloud Computing', 'Amazon Web Services (AWS)', 'Microsoft Azure', 'Google Cloud Platform (GCP)', 'Docker',
        'Kubernetes', 'Linux',
        'Shell Scripting', 'Cybersecurity', 'Network Security', 'Penetration Testing', 'Firewalls', 'Encryption',
        'Malware Analysis',
        'Digital Forensics', 'CI/CD', 'DevOps', 'Agile Methodology', 'Scrum', 'Kanban', 'Continuous Integration',
        'Continuous Deployment',
        'Software Development', 'Web Development', 'Mobile Development', 'Backend Development', 'Frontend Development',
        'Full-Stack Development',
        'UI/UX Design', 'Responsive Design', 'Wireframing', 'Prototyping', 'User Testing', 'Adobe Creative Suite',
        'Photoshop', 'Illustrator',
        'InDesign', 'Figma', 'Sketch', 'Zeplin', 'InVision', 'Product Management', 'Market Research',
        'Customer Development', 'Lean Startup',
        'Business Development', 'Sales', 'Marketing', 'Content Marketing', 'Social Media Marketing', 'Email Marketing',
        'SEO', 'SEM', 'PPC',
        'Google Analytics', 'Facebook Ads', 'LinkedIn Ads', 'Lead Generation', 'Customer Relationship Management (CRM)',
        'Salesforce',
        'HubSpot', 'Zendesk', 'Intercom', 'Customer Support', 'Technical Support', 'Troubleshooting',
        'Ticketing Systems', 'ServiceNow',
        'ITIL', 'Quality Assurance', 'Manual Testing', 'Automated Testing', 'Selenium', 'JUnit', 'Load Testing',
        'Performance Testing',
        'Regression Testing', 'Black Box Testing', 'White Box Testing', 'API Testing', 'Mobile Testing',
        'Usability Testing', 'Accessibility Testing',
        'Cross-Browser Testing', 'Agile Testing', 'User Acceptance Testing', 'Software Documentation',
        'Technical Writing', 'Copywriting',
        'Editing', 'Proofreading', 'Content Management Systems (CMS)', 'WordPress', 'Joomla', 'Drupal', 'Magento',
        'Shopify', 'E-commerce',
        'Payment Gateways', 'Inventory Management', 'Supply Chain Management', 'Logistics', 'Procurement',
        'ERP Systems', 'SAP', 'Oracle',
        'Microsoft Dynamics', 'Tableau', 'Power BI', 'QlikView', 'Looker', 'Data Warehousing', 'ETL',
        'Data Engineering', 'Data Governance',
        'Data Quality', 'Master Data Management', 'Predictive Analytics', 'Prescriptive Analytics',
        'Descriptive Analytics', 'Business Intelligence',
        'Dashboarding', 'Reporting', 'Data Mining', 'Web Scraping', 'API Integration', 'RESTful APIs', 'GraphQL',
        'SOAP', 'Microservices',
        'Serverless Architecture', 'Lambda Functions', 'Event-Driven Architecture', 'Message Queues', 'GraphQL',
        'Socket.io', 'WebSockets'
                     'Ruby', 'Ruby on Rails', 'PHP', 'Symfony', 'Laravel', 'CakePHP', 'Zend Framework', 'ASP.NET', 'C#',
        'VB.NET', 'ASP.NET MVC', 'Entity Framework',
        'Spring', 'Hibernate', 'Struts', 'Kotlin', 'Swift', 'Objective-C', 'iOS Development', 'Android Development',
        'Flutter', 'React Native', 'Ionic',
        'Mobile UI/UX Design', 'Material Design', 'SwiftUI', 'RxJava', 'RxSwift', 'Django', 'Flask', 'FastAPI',
        'Falcon', 'Tornado', 'WebSockets',
        'GraphQL', 'RESTful Web Services', 'SOAP', 'Microservices Architecture', 'Serverless Computing', 'AWS Lambda',
        'Google Cloud Functions',
        'Azure Functions', 'Server Administration', 'System Administration', 'Network Administration',
        'Database Administration', 'MySQL', 'PostgreSQL',
        'SQLite', 'Microsoft SQL Server', 'Oracle Database', 'NoSQL', 'MongoDB', 'Cassandra', 'Redis', 'Elasticsearch',
        'Firebase', 'Google Analytics',
        'Google Tag Manager', 'Adobe Analytics', 'Marketing Automation', 'Customer Data Platforms', 'Segment',
        'Salesforce Marketing Cloud', 'HubSpot CRM',
        'Zapier', 'IFTTT', 'Workflow Automation', 'Robotic Process Automation (RPA)', 'UI Automation',
        'Natural Language Generation (NLG)',
        'Virtual Reality (VR)', 'Augmented Reality (AR)', 'Mixed Reality (MR)', 'Unity', 'Unreal Engine', '3D Modeling',
        'Animation', 'Motion Graphics',
        'Game Design', 'Game Development', 'Level Design', 'Unity3D', 'Unreal Engine 4', 'Blender', 'Maya',
        'Adobe After Effects', 'Adobe Premiere Pro',
        'Final Cut Pro', 'Video Editing', 'Audio Editing', 'Sound Design', 'Music Production', 'Digital Marketing',
        'Content Strategy', 'Conversion Rate Optimization (CRO)',
        'A/B Testing', 'Customer Experience (CX)', 'User Experience (UX)', 'User Interface (UI)', 'Persona Development',
        'User Journey Mapping', 'Information Architecture (IA)',
        'Wireframing', 'Prototyping', 'Usability Testing', 'Accessibility Compliance', 'Internationalization (I18n)',
        'Localization (L10n)', 'Voice User Interface (VUI)',
        'Chatbots', 'Natural Language Understanding (NLU)', 'Speech Synthesis', 'Emotion Detection',
        'Sentiment Analysis', 'Image Recognition', 'Object Detection',
        'Facial Recognition', 'Gesture Recognition', 'Document Recognition', 'Fraud Detection',
        'Cyber Threat Intelligence', 'Security Information and Event Management (SIEM)',
        'Vulnerability Assessment', 'Incident Response', 'Forensic Analysis', 'Security Operations Center (SOC)',
        'Identity and Access Management (IAM)', 'Single Sign-On (SSO)',
        'Multi-Factor Authentication (MFA)', 'Blockchain', 'Cryptocurrency', 'Decentralized Finance (DeFi)',
        'Smart Contracts', 'Web3', 'Non-Fungible Tokens (NFTs)']


    skills = []

    for skill in skills_list:
        pattern = r"\b{}\b".format(re.escape(skill))
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            skills.append(skill)

    return skills


def extract_education_from_resume(text):
    # 1. Patterns for common degree prefixes
    # This catches "Bachelor of...", "Master of...", "Ph.D. in...", etc.
    degree_patterns = [
        r"\b(Bachelor|B\.?S\.?|B\.?A\.?|B\.?E\.?|B\.?Tech|B\.?C\.?A\.?)\b(?:\s+(?:of|in)\s+([A-Za-z\s&]+))?",
        r"\b(Master|M\.?S\.?|M\.?A\.?|M\.?E\.?|M\.?Tech|M\.?B\.?A\.?|M\.?C\.?A\.?)\b(?:\s+(?:of|in)\s+([A-Za-z\s&]+))?",
        r"\b(Ph\.?D\.?|Doctor of)\b(?:\s+(?:of|in)\s+([A-Za-z\s&]+))?",
        r"\b(Associate|Diploma|Certification|Secondary Education)\b(?:\s+(?:of|in)\s+([A-Za-z\s&]+))?"
    ]

    found_education = []

    for pattern in degree_patterns:
        # Find all matches in the text
        matches = re.finditer(pattern, text, re.IGNORECASE)
        for match in matches:
            # Clean the found string (remove extra whitespace/newlines)
            edu_str = match.group(0).strip()
            # Basic validation: ensure the matched string isn't too long (avoids false positives)
            if len(edu_str) < 60:
                found_education.append(edu_str)

    # 2. Catch specific stand-alone keywords for specialized fields
    broad_keywords = [
        'Computer Science', 'Information Technology', 'Nursing', 'Law',
        'Journalism', 'Psychology', 'Mechanical Engineering', 'Physics',
        'Chemistry', 'Biology', 'Economics', 'Finance', 'Accounting'
    ]

    for k in broad_keywords:
        if re.search(rf"\b{k}\b", text, re.IGNORECASE):
            # Only add if it's not already part of a degree string found above
            if not any(k.lower() in found.lower() for found in found_education):
                found_education.append(k)

    return list(set(found_education))

# ===================== ROUTES =====================
@app.route("/")
def home():
    return render_template("resume.html")

@app.route("/pred", methods=["POST"])
def pred():
    files = request.files.getlist("resumes[]")

    if not files or files[0].filename == "":
        return render_template("resume.html", message="No files uploaded")

    results = []

    for file in files:
        raw_text = pdf_to_text(file) if file.filename.endswith(".pdf") else file.read().decode("utf-8", errors="ignore")

        name = extract_name_from_resume(raw_text)
        email = extract_email_from_resume(raw_text)
        phone = extract_contact_number_from_resume(raw_text)

        cleaned = cleanResume(raw_text)
        category = predict_category(cleaned)
        job = job_recommendation(cleaned)
        skills = extract_skills_from_resume(cleaned)
        education = extract_education_from_resume(cleaned)
        result_data = {
            "name": name,
            "email": email,
            "phone": phone,
            "category": category,
            "job": job,
            "skills": skills,
            "education": education
        }

        save_resume_to_db(name, email, phone, category, job, skills, education)

        update_excel(result_data)

        results.append(result_data)

    return render_template("resume.html", results=results)

# ===================== RUN =====================
if __name__ == "__main__":
    init_db()
    app.run(debug=True)
